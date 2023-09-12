#使⽤两种⽅法打印出所有的⽔仙花数,所谓⽔仙花数是指⼀个三位数，其各位数字
#⽴⽅和等于该数本⾝。例如：153是⼀个⽔仙花数,因为 1³ + 5³ + 3³ 等于 153。
import random

x=input('please input 3d:')
if len(x) !=3:
    x=input('please input 3d:')
    if len(x)==3:
       bw=int(x)//100
       sw=int(x)//10%10
       gw=int(x)%10
       if bw**3+sw**3+gw**3==int(x):
           print('⽔仙花数')
       else:
           print('False')
'''
x=input('give number')
if 
'''

#输⼊⼀个⼩于1000的数字，产⽣对应的学号，例如：Python2023001、
#Python2023002、Python2023999。
str='Python2023'
re_str=''
a = input("please give number:")
if 0<int(a)<10:
    re_str=str+'00'+a
if 10<int(a)<100:
    re_str=str+'0'+a
if 100<int(a)<1000:
    re_str=str+a
print(re_str)

#企业发放的奖⾦根据利润提成。利润低于或等于10万元时，奖⾦可提10%；利润
# ⾼于10万元，低于20万元时，低于10万元的部分按10%提成，⾼于10万元的部
# 分，可提成7.5%；20万到40万之间时，⾼于20万元的部分，可提成5%；40万到
# 60万之间时⾼于40万元的部分，可提成3%；60万到100万之间时，⾼于60万元的
# 部分，可提成1.5%，⾼于100万元时，超过100万元的部分按1%提成，从键盘输
# ⼊当⽉利润I，求应发放奖⾦总数？
I=float(input('⽉利润:'))
if I <= 10:
    money = I * 0.1
elif 10 < I <= 20:
    money = 10 * 0.1 + (I - 10) * 0.075
elif 20 < I <= 40:
    money = 10 * 0.1 + 10 * 0.075 + (I - 20) * 0.05
elif 40 < I <= 60:
    money = 10 * 0.1 + 10 * 0.075 + 20 * 0.05 + (I - 40) * 0.03
elif 60 < I <= 100:
    money = 10 * 0.1 + 10 * 0.075 + 20 * 0.05 + 20 * 0.03 + (I - 60) * 0.015
else:
    money = 10 * 0.1 + 10 * 0.075 + 20 * 0.05 + 20 * 0.03 + 40 * 0.015 + (I - 100) * 0.01
print(money)

#产⽣⼀个⻓度为4位验证码，验证码涉及符号包括：数字、字⺟。
from itertools import product
list=[]
for i in range(0,10):
    list.append(str(i))
for j in range(ord('a'),ord('z')+1): #list.append(chr(ord(k)+32))
    list.append(chr(j))
for k in range(ord('A'),ord('Z')+1):
    list.append(chr(k))
print(len(list),list)

for le in range(4,5):
    result=product(list,repeat=le)
    for i in result:
        pwd=''.join(i)
        print(pwd)

#import random
#choice -->放回抽回n个
'''
import random
flag_list = []
for i in range(0, 10):
    flag_list.append(str(i))
for i in range(65, 91):
    flag_list.append(chr(i))
    flag_list.append(chr(i + 32))
print(flag_list)
# choices --> 放回的抽取N个
result = random.choices(flag_list, k=4)
print(result, ''.join(result))
print(random.choices([1, 2, 3], k=4))
'''

#这是经典的"百⻢百担"问题，有⼀百匹⻢，驮⼀百担货，⼤⻢驮3担，中⻢驮2
#担，两只⼩⻢驮1担，问有⼤，中，⼩⻢各⼏匹？（可以直接使⽤穷举法）
for dm in range(1,34):
    for zm in range(1,51):
        for xm in range(1,101):
            if dm*3+zm*2+xm/2==100 and dm+zm+xm==100:
                print(f'dm：{dm}，zm：{zm}，xm:{xm}')

#  使⽤openpyxl模块结合Python基础，编写⼀个登录注册系统：
# 同⼀个账号不能出现多次，注册时需要判断某个账户是否已存在。
# 登录时需要判断某个账号是否存在，再判断是否登录成功。
import openpyxl
wb = openpyxl.load_workbook('用户信息.xlsx')
ws1=wb.worksheets[0]
row = ws1.max_row
user=input('please give username')
password=input('please give password')
for i in range(1,row+1):
    if ws1.cell(i,1).value==user and ws1.cell(i,2).value==password:
         print('dlcg')
         wb.save('用户信息.xlsx')
         break
    if ws1.cell(i,1).value!=user:
        wb_1 = openpyxl.load_workbook('用户信息.xlsx')
        ws_1 = wb_1.worksheets[0]
        row = ws_1.max_row
        col = ws_1.max_column
        print('error ')
        new_user=input('please give username')
        new_password = input('please give password')
        ws_1.cell(row+1,1).value=new_user
        ws_1.cell(row+1,2).value=new_password
        wb_1.save('用户信息.xlsx')
        print('注册成功')
        break
    else:
        print('flase')



#判断⼀个数字是否是丑数（丑数定义详情⻅百度， 考验⾃学的时刻到了）
n = eval(input("输入："))
while n % 2 == 0:
    n = n / 2
while n % 3 == 0:
    n = n / 3
while n % 5 == 0:
    n = n / 5
if n == 1:
    print('True')
else:
    print('False')