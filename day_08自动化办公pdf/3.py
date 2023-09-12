import openpyxl
import xlrd
wb = openpyxl.load_workbook('用户信息.xlsx')
#row = ws1.max_row
#col = ws.max_column
ws1=wb.worksheets[0]
print()