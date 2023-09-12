import openpyxl
wb = openpyxl.load_workbook('用户信息.xlsx')
#row = ws1.max_row
#col = ws.max_column
ws1=wb.worksheets[0]
#print(ws1.cell(2,1).value)
#print(row)
row = ws1.max_row
user=input('please give username')
password=input('please give password')
for i in range(1,row+1):
    if ws1.cell(i,1).value==user and ws1.cell(i,2).value==password:
         print('dlcg')
         wb.save('用户信息.xlsx')
         break
    if ws1.cell(i,1).value!=user:
        wb_1 = openpyxl.load_workbook('用户信息.xlsx')
        ws_1 = wb_1.worksheets[0]
        row = ws_1.max_row
        col = ws_1.max_column
        #row = ws.max_row
        #col = ws.max_column
        #ws2 = wb.worksheets[0]
        print('error ')
        new_user=input('please give username')
        new_password = input('please give password')
        ws_1.cell(row+1,1).value=new_user
        ws_1.cell(row+1,2).value=new_password
        wb_1.save('用户信息.xlsx')
        print('注册成功')
        break
    else:
        print('flase')


#pass_word=input()

