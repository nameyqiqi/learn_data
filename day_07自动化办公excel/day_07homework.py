'''
import pandas as pd
import os
# 获取文件列表
file_list = os.listdir('数据集')
# 创建空的DataFrame对象
merged_data = pd.DataFrame()
# 循环读取文件并合并数据
for file_name in file_list:
    file_path = '数据集/' + file_name
    data = pd.read_excel(file_path)
    merged_data = pd.concat([merged_data, data])
# 保存合并后的数据
merged_data.to_excel('merged.xlsx', index=False)
'''

import openpyxl
import tqdm

# 6. 新建一个新的合集文件
new_wb = openpyxl.Workbook()
# openpyxl模块默认生成一个名为sheet的表

# 1. 构造文件路径
area = ['大邑', '简阳', '都江堰', '金牛', '高新', '高新西']
# 进度条tqdm方法操作对象是容器
for i in tqdm.tqdm(area):
    file_path = './数据集/成都' + i + '二手房信息.xlsx'
    # 2. 加载某文件
    wb = openpyxl.load_workbook(file_path)
    # 3. 查看已存在的工作表（sheet）
    ws_list = wb.worksheets
    # print(ws_list)
    # 4. 指向表
    ws = ws_list[0]
    # 5. 查看最大行列数：max_row、max_column
    rows = ws.max_row
    cols = ws.max_column

    # 7. 在new_wb中，每个城市的数据各自一张表
    new_wb.create_sheet(i)
    new_ws = new_wb[i]
    # 8. 写入数据
    # 获取行
    for row in range(1, rows + 1):
        # 获取列
        for col in range(1, cols + 1):
            # 给新的表的某单元格填入旧表的对应单元格的数据
            new_ws.cell(row, col).value = ws.cell(row, col).value

# 删除xlsx文件中的工作表：
# 工作簿.remove(工作表): 此处的工作表必须先指向
new_wb.remove(new_wb['Sheet'])

# 9. 保存关闭
new_wb.save('./成都二手房数据合集.xlsx')
print('写入完成')