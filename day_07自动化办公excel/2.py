# 二、成都二手房合集压缩包中的所有xlsx文件合并到一个excel文件中，可以一个工作表存放一个行政区的数据。
# 三、复习前面讲过的知识点 (尝试记忆每个方法及功能
import openpyxl, tqdm
suo_1 = openpyxl.Workbook()
suo =['大邑','简阳','都江堰','金牛','高新','高新西']
for i in suo:
    seo='./数据集/成都'+i+'二手房信息.xlsx'
    wb = openpyxl.load_workbook(seo)
    sb = wb.worksheets
    st = sb[0]
    h = st.max_row
    l = st.max_column
    suo_1.create_sheet(i)
    print(h ,l)
    sb_1 = suo_1[i]
    for i in range(1,h+1):
        for j in range(1,l+1):
           sb_1.cell(i,j).value=st.cell(i,j).value
suo_1.save('./数据集/d.xlsx')