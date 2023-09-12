import openpyxl

wb = openpyxl.load_workbook('./第七次全国人口普查数据.xlsx')
ws_list = wb.worksheets
print(ws_list)
ws = ws_list[0]
rows = ws.max_row
# x轴刻度标签必须是字符串
# 表示数据的y轴必须是数字
city = []
man = []
woman = []
for i in range(3, rows + 1):
    city.append(ws.cell(i, 1).value)
    man.append(ws.cell(i, 2).value)
    woman.append(ws.cell(i, 3).value)
print(city, man, woman)

from pyecharts.charts import Bar
from pyecharts import options as opts

A = (
    Bar(init_opts=opts.InitOpts(width='1300px', height='800px'))
    .add_xaxis(xaxis_data=city)
    .add_yaxis(series_name='男', y_axis=man, stack="one")
    .add_yaxis(series_name='女', y_axis=woman, stack="one")
    .reversal_axis()
    .set_global_opts(
        yaxis_opts=opts.AxisOpts(
            axislabel_opts=opts.LabelOpts(interval=0)
        ),
        datazoom_opts=opts.DataZoomOpts(orient='vertical')
    )
    .set_series_opts(
        label_opts=opts.LabelOpts(formatter='{c}%')
    )
    .render('各省份性别比例.html')
)