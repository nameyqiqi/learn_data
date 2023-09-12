import openpyxl
wb = openpyxl.load_workbook('第七次全国人口普查数据.xlsx')
ws_list = wb.worksheets
print(ws_list)
ws = ws_list[0]
rows = ws.max_row
col = ws.max_column
area_list = []
ratio_list = []
m_list = []
wo_list = []
for i in range(2,rows+1):
    area=ws.cell(i,1).value
    ratio=ws.cell(i,4).value
    man=ws.cell(i,2).value
    wo=ws.cell(i,3).value
    area_list.append(area)
    ratio_list.append(ratio)
    m_list.append(man)
    wo_list.append(wo)
print(area_list,ratio_list)
# 折线图
import pyecharts.options as opts
from pyecharts.charts import Line
c = (
    Line()
    .add_xaxis(area_list)
    .add_yaxis("地区",ratio_list)
    #.add_yaxis("商家B", Faker.values())
    .set_global_opts(
        title_opts=opts.TitleOpts(title="男女比例"),
        datazoom_opts=opts.DataZoomOpts(orient='horizontal'),
                     )
    .render("男女比例.html")
)
# 柱状图
from pyecharts import options as opts
from pyecharts.charts import Bar

d = (
    Bar()
    .add_xaxis(area_list)
    .add_yaxis("男", m_list,stack='pe')
    .add_yaxis("女", wo_list,stack='pe')
    .reversal_axis()
    .set_global_opts(
        title_opts=opts.TitleOpts(title="地区男女比例", subtitle=""),
        datazoom_opts=opts.DataZoomOpts(orient='vertical'),

    )
    .set_series_opts(
        markpoint_opts=opts.MarkPointOpts(
        )
    )
    .render("地区男女比例.html")
)

#人口占比
from pyecharts.charts import Pie
ws = ws_list[1]
rows=ws.max_row
pe_list=[]
#cols=ws.max_column
for i in range(3,rows+1):
    age=ws.cell(i,1).value
    pe=ws.cell(i,2).value
    pe_list.append([age,pe])
print(pe_list)
A = (

    Pie()
    .add(series_name="", data_pair=pe_list)
    # 引入全局配置项
    .set_global_opts(
        # 从全局配置项中引入标题配置项,
        # 配置项等号左边怎么写：配置项名字字母全小写，opts前加下划线
        # title:主标题内容    pos_left:标题距离容器左侧的距离（position）
        title_opts=opts.TitleOpts(title='人口占比', pos_left='center'),
        # 添加图例配置项: orient：改图例的水平或垂直布局，pos_top/left修改图例距离上方/左侧的距离
        legend_opts=opts.LegendOpts(is_show=True, pos_top='7%', pos_left='right', orient='vertical'),
        # 添加提示框配置项：通过formatter修改提示框内容样式，画饼图时，formatter有一个{d}表示占比
        # {b}:数据名, {c}:数值， {d}占比
        tooltip_opts=opts.TooltipOpts(formatter='{b}：{c}（{d}%）')
    )
    # 生成图表
    .render("./人口占比.html")
)
ws = ws_list[2]
rows = ws.max_row
a_list=[]
b_list=[]
c_list=[]
d_list=[]
for i in range(2,rows+1):
    a=ws.cell(i,2).value
    b=ws.cell(i,3).value
    c=ws.cell(i,4).value
    d=ws.cell(i,5).value
    a_list.append(a)
    b_list.append(b)
    c_list.append(c)
    d_list.append(d)
c = (
    Bar(init_opts=opts.InitOpts(width='1300px', height='800px'))
    .add_xaxis(area_list)
    .add_yaxis("0~14岁", a_list,stack='one')
    .add_yaxis("15~59岁", b_list,stack='one')
    .add_yaxis("60岁以上", c_list,stack='one')
    .add_yaxis("65岁以上", d_list,stack='one')

    # 柱状图转条形图:reversal_axis()
    .reversal_axis()
    # set_global_opts:全局配置项
    .set_global_opts(
        title_opts=opts.TitleOpts(title="年龄构成", subtitle="地区"),
        # 引入区域缩放配置项, is_show默认为True
        # 区域缩放配置项默认水平排列，当柱状图变为条形图，可能需要区域缩放跟着一起变
        # orient=horizontal(水平，默认)，vertical（垂直）
        datazoom_opts=opts.DataZoomOpts(orient='vertical'),
        # 引入坐标轴配置项, 如果是修改坐标轴，一定一定说明是哪个轴
        yaxis_opts=opts.AxisOpts(
            # 通过axislabel_opts参数，得知，再引入标签配置项
            # # 引入标签配置项修改坐标轴刻度的显示间隔，interval改为0，强制显示所有
            # rotate:标签旋转
            axislabel_opts=opts.LabelOpts(interval=0, rotate=-7)
        )
    )
    # set_series_opts:系列配置项
    # 如何指定到修改x轴刻度标签：坐标轴 > x坐标轴标签 > 标签
    # .set_series_opts(
    #     # 引入标签配置项修改坐标轴刻度的显示间隔，interval改为0，强制显示所有
    #     label_opts=opts.LabelOpts(interval=0)
    # )
    .render("年龄构成.html")
)
from pyecharts import options as opts
from pyecharts.charts import Map
ws = ws_list[2]
rows = ws.max_row
old_minlist=[]
old_list=[]
for k in range(2,rows+1):
    old_pe=[ws.cell(k,1).value,ws.cell(k,5).value]
    old_list.append(old_pe)

c = (
    Map()
    .add(series_name="", data_pair=old_list, maptype="china")
    .set_global_opts(
        title_opts=opts.TitleOpts(title="老龄"),
        # 添加视觉映射配置项
        visualmap_opts=opts.VisualMapOpts(min_=5.67, max_=17.24),
        toolbox_opts=opts.ToolboxOpts()

    )
    .render("地区老龄化.html")
)

