import openpyxl

wb = openpyxl.load_workbook('第七次全国人口普查数据.xlsx')
ws_list = wb.worksheets
print(ws_list)
ws = ws_list[-2]

# 问题一：分析全国各省人口：柱状图
city = []
people = []
rows = ws.max_row
# 构造柱状图需要的数据
for i in range(3, rows):
    city.append(ws.cell(i, 1).value)
    num = ws.cell(i, 2).value / 10000000
    # round(数值， 小数位数)：四舍五入保留小数
    people.append(round(num, 2))
print(city, people)

# 画柱状图
from pyecharts.charts import Bar
from pyecharts import options as opts

A = (
    Bar()
    .add_xaxis(xaxis_data=city)
    .add_yaxis(series_name="", y_axis=people)
    .set_global_opts(
        # 引入坐标轴配置项修改y轴的名称
        yaxis_opts=opts.AxisOpts(name='人口数（单位：千万）'),
        # 引入图例配置项,一个图没有图例，默认图例的灰色方块依旧存在
        legend_opts=opts.LegendOpts(is_show=False),
        # 如果涉及柱子很多，加上区域缩放效果会好一些
        datazoom_opts=opts.DataZoomOpts(is_show=True),
        # 默认x轴刻度标签全显示
        xaxis_opts=opts.AxisOpts(
            # 在引入对标签的修改
            axislabel_opts=opts.LabelOpts(interval=0, rotate=-30)
        ),
        # 标题配置项
        title_opts=opts.TitleOpts(title='各省人口数量普查', pos_left='center')
    )
    .render('./各省人口数量.html')
)

# 问题二：将10-20年各省份人口数量变化反馈到地图上
# pyecharts的地图需要的数据必须是：北京市、重庆市、四川省等
print(city)

new_city = ['北  京 市', '天  津 市', '河  北 省', '山  西 省', '内蒙古 自治区',
            '辽  宁 省', '吉  林 省', '黑龙江 省', '上  海 市', '江  苏 省',
            '浙  江 省', '安  徽 省', '福  建 省', '江  西 省', '山  东 省',
            '河  南 省', '湖  北 省', '湖  南 省', '广  东 省', '广  西 壮族自治区',
            '海  南 省', '重  庆 市', '四  川 省', '贵  州 省', '云  南 省',
            '西  藏 自治区', '陕  西 省', '甘  肃 省', '青  海 省', '宁  夏 回族自治区',
            '新  疆 维吾尔自治区']

# 地图需要的数据：[[北京市, 1], [上海市, 1.2]]
for i in range(0, len(city)):
    for j in new_city:
        # 如果city中拿到的元素与j处理完空格之后按照city拿到的元素长度切片的前几位相等
        sample = j.replace(' ', '')
        change = city[i].replace(' ', '')
        if change == sample[0:len(change)]:
            # 修改列表中元素
            city[i] = sample
            break
print(city)

# 计算20年相对于10年省份人口比例变化情况
people_change = []
for i in range(3, rows):
    people_change.append(ws.cell(i, 3).value - ws.cell(i, 4).value)
print(people_change)


data = []
for i in range(0, len(city)):
    # [city[i], people[i]]
    data.append([city[i], people_change[i]])

from pyecharts.charts import Map
from pyecharts import options as opts
B = (
    Map()
    .add(series_name="", data_pair=data)
    .set_global_opts(
        visualmap_opts=opts.VisualMapOpts(
            min_=min(people_change), max_=max(people_change)
        )
    )
    .render('./各省份20-10年分口比例变化.html')
)