from pyecharts import options as opts
from pyecharts.charts import Map
import openpyxl
from pyecharts.charts import Line
wb = openpyxl.load_workbook('第七次全国人口普查数据.xlsx')
ws_list = wb.worksheets
print(ws_list)
ws = ws_list[2]
rows = ws.max_row
old_list=[['北京市', 13.3], ['天津市', 14.75], ['河北省', 13.92], ['山西省', 12.9],
          ['内蒙古自治区', 13.05], ['辽宁省', 17.42], ['吉林省', 15.61], ['黑龙江省', 15.61],
          ['上海市', 16.28], ['江苏省', 16.2], ['浙江省', 13.27], ['安徽省', 15.01],
          ['福建省', 11.1], ['江西省', 11.89], ['山东省', 15.13], ['河南省', 13.49],
          ['湖北省', 14.59], ['湖南省', 14.81], ['广东省', 8.58], ['广西壮族自治区', 12.2],
          ['海南省', 10.43], ['重庆市', 17.08], ['四川省', 16.93], ['贵州省', 11.56],
          ['云南省', 10.75], ['西藏自治区', 5.67], ['陕西省', 13.32], ['甘肃省', 12.58],
          ['青海省', 8.68], ['宁夏回族自治区', 9.62], ['新疆维吾尔自治区', 7.76]]
list=[]
for k in range(2,rows+1):
    if ws.cell(k-1,13).vlaue !=ws.cell(k,13).vlaue:
        list.append(k)
        print(k)