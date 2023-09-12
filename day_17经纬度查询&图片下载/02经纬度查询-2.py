import openpyxl
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
# 驱动打开浏览器
browser = webdriver.Chrome()
browser.get('https://map.yanue.net/')
time.sleep(2)


# 把经纬度.xlsx 文件中的 第一列 和第二列 进行组合    经度，纬度
# 1. 打开已有文件
wb = openpyxl.load_workbook('./经纬度.xlsx')
# 2.选择工作表    工作簿.active  活跃的 不一定是想要的
ws = wb.active
# 3. 获取工作表中的最大行数
max_row = ws.max_row
# 4. 遍历每一行数据
for i in range(2,max_row+1):
    jd = ws.cell(i,1).value
    wd = ws.cell(i,2).value
    # print(jd,wd,type(jd),type(wd))
    data = f'{jd},{wd}'
    # 定位输入框   输入框中内容清空
    input_tag = browser.find_element(by=By.CSS_SELECTOR, value='#latLng')
    input_tag.clear()
    # 输入新内容  data
    input_tag.send_keys(data)
    time.sleep(0.5)
    # 点击 解析经纬度
    button_tag = browser.find_element(by=By.CSS_SELECTOR, value='#toAddressBtn')
    button_tag.click()
    time.sleep(1)
    # 获取解析的结果
    res = browser.find_element(by=By.CSS_SELECTOR, value='#showResults').text
    # 结果处理  只要 地址
    adrr = res.split('：')[-1]
    # 地址 写入到excel文件     行  i   列  3
    ws.cell(i,3).value = adrr
    # 保存 excel
    wb.save('./经纬度.xlsx')

browser.close()








