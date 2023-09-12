import openpyxl
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
# 驱动打开浏览器
browser = webdriver.Chrome()
browser.get('https://map.yanue.net/')
time.sleep(2)

# 把经纬度.xlsx 文件中的 第一列 和第二列 进行组合    经度，纬度
# 1. 打开已有文件
wb = openpyxl.load_workbook('./经纬度.xlsx')
# 2.选择工作表    工作簿.active  活跃的 不一定是想要的
ws = wb.active
# 3. 获取工作表中的最大行数
max_row = ws.max_row
# 4. 遍历每一行数据
for i in range(2,max_row+1):
    jd = ws.cell(i,5).value
    wd = ws.cell(i,6).value
    # print(jd,wd,type(jd),type(wd))
    data = f'{jd}{wd}'
    #print(data)
    input_tag=(browser.find_element(By.XPATH,'//*[@id="addr"]'))
    input_tag.clear()
    time.sleep(0.5)
    input_tag.send_keys(data)
    time.sleep(1)
    #点击查询
    sub_tag=browser.find_element(By.XPATH,'//*[@id="toLatLngBtn"]').click()
    time.sleep(1)
    # 获取解析的结果
    res = browser.find_element(by=By.CSS_SELECTOR, value='#showResults').text
    # 结果处理  只要 经纬度
    adrr = res.split(':')[-1]
    print(adrr)
    #地址 写入到excel文件     行  i   列  3
    ws.cell(i, 7).value = adrr
    # 保存 excel
    wb.save('./经纬度.xlsx')
