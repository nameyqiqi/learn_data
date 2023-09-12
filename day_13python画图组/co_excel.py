import openpyxl
wb = openpyxl.load_workbook('./tu/中国新冠疫情数据.xlsx')
ws = wb.worksheets[0]
rows = ws.max_row
cols = ws.max_column
#print(rows,cols)#5025 13
city=['天津市', '山东省', '湖北省', '上海市', '北京市', '广西壮族自治区', '广东省',
      '甘肃省', '山西省', '四川省', '河北省', '台湾', '香港', '辽宁省', '云南省',
      '湖南省', '澳门', '河南省', '黑龙江省', '重庆市', '吉林省', '青海省', '浙江省',
      '江苏省', '陕西省', '新疆维吾尔自治区', '内蒙古自治区', '西藏自治区', '贵州省',
      '江西省', '福建省', '安徽省', '宁夏回族自治区', '海南省']
list =[] #不同元素下标
for j in range(2,rows):
    if ws.cell(j,13).value!=ws.cell(j+1,13).value:
        list.append(j)
print(list)