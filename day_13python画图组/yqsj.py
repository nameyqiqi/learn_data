import openpyxl

wb = openpyxl.load_workbook('./tu/中国新冠疫情数据.xlsx')
ws = wb.worksheets[0]
#1已确定 2增加 3固化 4固化incr 5目前已确定 6目前增加 7死亡日期 8死亡记数 9死亡增加量 10可疑 11可疑增 12省份
rows = ws.max_row
city=['天津市', '山东省', '湖北省', '上海市', '北京市', '广西壮族自治区', '广东省',
      '甘肃省', '山西省', '四川省', '河北省', '台湾', '香港', '辽宁省', '云南省',
      '湖南省', '澳门', '河南省', '黑龙江省', '重庆市', '吉林省', '青海省', '浙江省',
      '江苏省', '陕西省', '新疆维吾尔自治区', '内蒙古自治区', '西藏自治区', '贵州省',
      '江西省', '福建省', '安徽省', '宁夏回族自治区', '海南省']
p_list =[]
tlist=[]
# die_total=0
# for i in range(2,rows+1):
#     city.append(ws.cell(i,13).value)
# print(set(city))
from pyecharts.charts import Map
from pyecharts import options as opts
for j in range(2,rows+1):
    for i in range(0,len(city)):
        if ws.cell(j,13).value==city[i]:
            p_list.append([ws.cell(j,9).value+ws.cell(j,10).value])
            break
new_plist=[]
print(p_list)

# c = (
#     Map()
#     .add(series_name="", data_pair=p_list, maptype="china")
#     .set_global_opts(
#         title_opts=opts.TitleOpts(title="Map-基本示例"),
#         # 添加视觉映射配置项
#         visualmap_opts=opts.VisualMapOpts(min_=50, max_=500),
#         toolbox_opts=opts.ToolboxOpts()
#
#     )
#     .render("map_base.html")
# )
#
# print(p_list)




