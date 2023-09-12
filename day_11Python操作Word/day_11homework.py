#作业：
#1. 复习整理笔记（以后做笔记可以用语雀，毕竟多端同步，有望就能学）
#2. 使用各地区性别构成.xlsx画一个全国男女比例饼图
#3. 使用各地区人口.xlsx画一个柱状图展示各地区人口数
# 使用各地区性别构成.xlsx画一个全国男女比例饼图
# import openpyxl
#
# # 1. 加载文件
# wb = openpyxl.load_workbook('./各地区性别构成.xlsx')
# # 2. 查看所有工作表，通过下标指定
# ws = wb.worksheets[0]
# # print(ws)
# # 3. 读取数据，构造成饼图需要的数据样式
# data = [['男', ws.cell(2, 2).value], ['女', ws.cell(2, 3).value]]
# print(data)
# # --------------------------------------------------------------
# from pyecharts.charts import Pie
# from pyecharts import options as opts
#
# # 链式调用
# A = (
#     Pie()
#     .add(series_name="", data_pair=data)
#     .set_global_opts(
#         # 引入提示框配置项
#         tooltip_opts=opts.TooltipOpts(formatter='{b}：{d}%')
#     )
#     # 生成图表
#     .render('./全国人口性别饼图.html')
# )


import openpyxl
from pyecharts.charts import Pie
from pyecharts import options as opts
# 导入主题库
from pyecharts.globals import ThemeType

wb = openpyxl.load_workbook('./图表/各地区性别构成.xlsx')
ws = wb.worksheets[0]
rows = ws.max_row
#print(rows)
list_xb = []
list_rk_x = []
list_rk_y =[]
#男女比例data
for i in range(2,rows+1):
    result = [ws.cell(i,1).value,ws.cell(i,4).value]
    list_xb.append(result)
#print(list_xb)
#各地区人口数据
wb_1 = openpyxl.load_workbook('./图表/各地区人口.xlsx')
ws_1 = wb_1.worksheets[0]
p_rows = ws_1.max_row
for j in range(3,p_rows-1):
    result_p_xxios = ws_1.cell(j,1).value
    result_p_yaxios= ws_1.cell(j,2).value
    list_rk_x.append(result_p_xxios)
    list_rk_y.append(result_p_yaxios)
print(list_rk_x)
print(list_rk_y)
#print(list_rk)
A = (
    # 声明饼图：初始化是在声明图形时就规定好的
    Pie(init_opts=opts.InitOpts(theme=ThemeType.CHALK))
    # 向饼图添加数据
    # 系列名称(series_name)属于大分类、这个图的数据对应的内容是小分类,系列名称如果没有设为空字符串
    # 第二个参数为图表需要的数据，饼图：[[名称(小分类), 数值],[名称(小分类), 数值],[名称(小分类), 数值]]
    # radius:饼图的内外半径（设置环形图的关键）
    .add(series_name="", data_pair=list_xb, radius=['50%', '70%'])
    # 引入全局配置项
    .set_global_opts(
        # 从全局配置项中引入标题配置项,
        # 配置项等号左边怎么写：配置项名字字母全小写，opts前加下划线
        # title:主标题内容    pos_left:标题距离容器左侧的距离（position）
        title_opts=opts.TitleOpts(title='男女比例', pos_left='center'),
        # 添加图例配置项: orient：改图例的水平或垂直布局，pos_top/left修改图例距离上方/左侧的距离
        legend_opts=opts.LegendOpts(is_show=True, pos_top='7%', pos_left='right', orient='vertical'),
        # 添加提示框配置项：通过formatter修改提示框内容样式，画饼图时，formatter有一个{d}表示占比
        # {b}:数据名, {c}:数值， {d}占比
        tooltip_opts=opts.TooltipOpts(formatter='{b}：{c}（{d}%）')
    )
    # 生成图表
    .render("./图表/男女比例.html")
)
from pyecharts.charts import Bar
B = (
     Bar()
    .add_xaxis(list_rk_x)
    .add_yaxis("地区", list_rk_y)
    .set_global_opts(
         datazoom_opts=opts.DataZoomOpts(orient='vertical'),
         title_opts=opts.TitleOpts(title="地区人口数", subtitle="地区人口"),
         #datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_="inside")],
         xaxis_opts=opts.AxisOpts(
            # 通过axislabel_opts参数，得知，再引入标签配置项
            # # 引入标签配置项修改坐标轴刻度的显示间隔，interval改为0，强制显示所有
            # rotate:标签旋转
            axislabel_opts=opts.LabelOpts(interval=0, rotate=-7)
        )
         )

    .reversal_axis()
    .render("地区人口.html")
)