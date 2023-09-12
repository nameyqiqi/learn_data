from pyecharts import options as opts
from pyecharts.charts import Map
import openpyxl
wb = openpyxl.load_workbook('第七次全国人口普查数据.xlsx')
ws_list = wb.worksheets
print(ws_list)
ws = ws_list[3]
rows = ws.max_row
old_list=[]
for k in range(2,rows+1):
    old_pe=[ws.cell(k,1).value,ws.cell(k,5).value]
    old_list.append(old_pe)
print(old_list)
c = (
    Map()
    .add(series_name="", data_pair=old_list, maptype="china")
    .set_global_opts(
        title_opts=opts.TitleOpts(title="老龄"),
        # 添加视觉映射配置项
        visualmap_opts=opts.VisualMapOpts(min_=5.67, max_=17.24),
        toolbox_opts=opts.ToolboxOpts()

    )
    .render("地区老龄化.html")
)
