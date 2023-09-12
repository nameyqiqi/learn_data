# 1. 读取一个已存在的文件： load_workbook('文件路径与文件名')
import openpyxl
wb = openpyxl.load_workbook('./数据集/邮政编码.xlsx')
# 2. worksheets： 查看已存在的所有工作表
ws_list = wb.worksheets
print(ws_list)
# 3. 指定工作表
ws = wb['邮政编码']
print(ws)
# 可以通过下标从所有的工作表（worksheets）中指定其中某个表
ws_1 = ws_list[0]
print(ws == ws_1)

# 4. 查看有多少行、列
rows = ws.max_row
columns = ws.max_column
print(f'一共{rows}行{columns}列')

# 5. 获取每个单元格的内容
for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        result = ws.cell(i, j).value
        print(result, end='\t')
    print()