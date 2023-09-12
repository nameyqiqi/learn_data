'''
import openpyxl
wb = openpyxl.load_workbook('./数据集/北京高档酒店价格分析.xlsx')
ws = wb['北京高档酒店价格分析']
row = ws.max_row
col = ws.max_column
for i in range(2,row+1):
    total = 0
    for j in range(4,8):
        total+=float(ws.cell(i,j).value)
        ws.cell(1,17).value='average'
        ws.cell(i,17).value=total/4
        round(total/4,1)
    #print(total)
wb.save('1.xlsx')
'''
'''
import openpyxl
wb = openpyxl.load_workbook('./数据集/北京高档酒店价格分析.xlsx')
# 指定工作表
ws = wb.worksheets[0]
rows = ws.max_row
# 外层循环获取每一个酒店信息
for i in range(2, rows + 1):
    # total:每次拿出一个酒店的信息，先清零，再累加
    total = 0
    for j in range(4, 8):
        result = ws.cell(i, j).value
        # print(result, type(result))
        total += result
    name = ws.cell(i, 1).value
    # round(数值, 小数位数)：四舍五入保留小数位数
    print(f'{name}的平均评分为{round(total / 4, 1)}')
'''
import openpyxl
# 三、计算各区二手房平均单价，再写入到文件中

wb = openpyxl.load_workbook('./数据集/成都二手房.xlsx')
# 查看所有的工作表得到的是列表，列表中每个元素是每一个工作表
ws_list = wb.worksheets
# 用一个列表存储所有行政区的平均单价
data = []
# 使用循环分别指定每一个表
for st in ws_list:
    # 创建一个列表，为了求和与求平均，每换一个行政区，清空
    unit_list = []
    # 循环一次，获取这一张表的行数
    rows = st.max_row
    for i in range(2, rows + 1):
        result = st.cell(i, 8).value
        # NoneType空值类型中唯一一个元素None
        if result != None:
            # 指定需要替换掉的符号
            replace_flag = ['参考价:', ',', '元/平']
            # 循环一次，替换一个符号，没有的符号自动跳过，将替换的结果
            # 重新赋给result进入下一次循环继续进行替换
            for i in replace_flag:
                result = result.replace(i, '')
            # print(result)
            unit_list.append(int(result))
    name = st.cell(2, 1).value
    unit_price = sum(unit_list) / len(unit_list)
    # round:如果小数部分末尾为0，自动忽略
    print(f'{name}的二手房平均单价{round(unit_price, 2)}')
    data.append([name, unit_price])
print(data)
# -----计算每个行政区的单价-----
# -----将平均单价写入新的工作表-----
sheet_name = '各区均价'
wb.create_sheet(sheet_name)
ws = wb[sheet_name]
for i in range(1, len(data) + 1):
    # i - 1， j - 1， 两层for循环趋向于了excel的行列的初始从1开始，
    # 列表的下标从0开始
    for j in range(1, len(data[i - 1]) + 1):
        # 通过行列号获取某个位置需要的元素，写入到单元格
        ws.cell(i, j).value = (data[i - 1])[j - 1]

# 保存关闭
# 先用另存为的方式，保证打开的与保存的路径不一样。
# 先不用源文件保存的方式，会遇到一个错误
wb.save('./数据集/成都链家二手房.xlsx')







