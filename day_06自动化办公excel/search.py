import openpyxl
wb = openpyxl.load_workbook('./数据集/邮政编码.xlsx')
#指定工作表
ws_list=wb.worksheets
ws=ws_list[0]
print(ws_list)
#获取第一行，判断根据某一列获取另一列
cols=ws.max_column
for i in range(1,cols+1):
    result=ws.cell(1,i).value
    print(result)
rows=ws.max_row
flag=False
area=input('please input:')
for i in range(2,rows+1):
    if area==ws.cell(i,2).value:
        print(ws.cell(i,1).value)
        flag=True
        break
if flag==False:
    print('未找到')










