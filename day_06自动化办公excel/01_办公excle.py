# 一、Excel后缀名
# .xls和.xlsx
# .xls是excel 2007以前； .xlsx是excel 2007及以后

# 二、两种后缀名区别
# 1.
# .xls: 上限最多存储(65536行, 256列)数据
# .xlsx: 上限最多存储(1048576行, 16384列)数据

# 2. 两种后缀名文件，存储同样的数据，xlsx占内存小，处理速度快。

# 三、办公软件：office、wps、永中office

# 四、python操作Excel
# xlrd、xlwt：针对于xls文件进行读写操作的。
# openpyxl：针对于xlsx文件进行读写操作。

# 五、python操作文件的顺序与手动操作顺序差不多

# 六、一个Excel文件是一个工作簿（workbook）。
# 每一个小方格叫做单元格（cell）。
# 工作簿与单元格之间还有一个页面，叫做工作表(worksheet)。

# 七、安装模块
# 导包方式
# 1. import 模块名
# 2. from 模块名 import 方法名

import openpyxl

# 1. 创建并打开文件
wb = openpyxl.Workbook()

# 2. 直接写入内容放入单元格（cell）
# 要求，需要告知写入哪个表的哪个单元格
ws = wb['Sheet']    # 指定工作表

# 如果是cell方法，行列都是从1开始，cell(行, 列)
# 需要使用value，表示进行值的操作
ws.cell(1, 1).value = '张三'

# 按照标准的行和列指定单元格
ws['AA2'].value='李四'
ws['A10'].value='wsa'

# 创建新的工作表create_sheet
wb.create_sheet('sheet1')
wb.create_sheet('sheet2',1)

# 查看某个工作簿的所有工作表
ws_list=wb.worksheets
print(ws_list)

# max_row max_column :查看最大行和列
print(ws.max_row,ws.max_column)

# 2. finally:保存、关闭、设置存储路径、重命名文件
wb.save('学生成绩.xlsx')
# 一、创建Excel工作簿：Workbook()
# 二、指定某个工作表:  工作簿[工作表名]
# 三、自行创建工作表： create_sheet(工作表名, 工作表的位置(下标))
#      工作表的位置(下标)如果不写，默认最后追加
# 四、操作单元格
# a. 标准的行列号的说法： 工作表[列号行号]
# b. 使用cell指定行列号：  工作表.cell(行, 列)
# 一定要使用value声明操作的是值
# 五、保存关闭：工作簿.save('文件名.xlsx')