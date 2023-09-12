import openpyxl
import random

# 创建、打开工作簿
wb = openpyxl.Workbook()
# 创建工作簿默认存在一个Sheet
ws = wb['Sheet']
# 写入数据
sheet_title = ['姓名', '语文', '数学', '英语']
# openpyxl模块操作的xlsx文件行列均从1开始
for i in range(0, len(sheet_title)):
    ws.cell(1, i + 1).value = sheet_title[i]

names = ['黄忠', '关羽', '赵云', '马超', '张飞']
# 此循环控制数据条数
for i in range(0, len(names)):
    # 写的是姓名
    ws.cell(i + 2, 1).value = names[i]
    # 此循环写成绩
    for j in range(2, 5):
        ws.cell(i + 2, j).value = random.randint(0, 100)

wb.save('小学成绩表.xlsx')
print('写入完成')