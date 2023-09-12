import openpyxl
wb=openpyxl.load_workbook('./数据集/成都二手房.xlsx')
ws_list=wb.worksheets
#print(ws_list)
list=[]
data=[]
for st in ws_list:
    rows=st.max_row
    for i in range(2,rows+1):
        replace_list=['参考价:', ',', '元/平']
        result=st.cell(i,8).value
        if result !=None:
            for j in replace_list:
                result=result.replace(j,'')
            #print(result)
            #print(type(result))
            name=st.cell(2,1).value
            list.append(int(result))
    average=sum(list)/len(list)
    data.append([name,average])
print(data)
wb.create_sheet('各区均价')
ws = wb['各区均价']
for i in range(1,len(data)+1):
    for j in range(1,len(data[i-1])+1):
        ws.cell(i,j).value=data[i-1][j-1]
wb.save('11.xlsx')

